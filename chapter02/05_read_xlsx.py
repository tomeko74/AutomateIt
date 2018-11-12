# Get all the sheets from given workbook
import openpyxl

workbook = openpyxl.load_workbook('myxls.xlsx')
print("Workbook Object:", workbook.sheetnames)

# Get a particular sheet of the workbook
people = workbook['People']
print("People sheet object:", people)

# Get individual cell objects with cell name or column/row combination
import openpyxl

workbook = openpyxl.load_workbook('myxls.xlsx')
people = workbook['People']
print("First cell Object:", people['A1'])
print("Other Cell Object:", people.cell(row=3, column=2))

# Read cell values from cell objects
print("First Name:", people['B2'].value, people['C2'].value)

